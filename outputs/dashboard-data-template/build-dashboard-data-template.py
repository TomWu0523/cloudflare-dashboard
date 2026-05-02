import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

output_dir = Path(__file__).resolve().parent
project_root = output_dir.parents[1]
dashboards = json.loads((project_root / "data" / "dashboard-data.json").read_text(encoding="utf-8"))

product_labels = {
    "magnus1180": "1180 Magnus OR Table",
    "tegris": "Tegris Dashboard",
    "icMic": "IC MIC Dashboard",
}

headers = [
    "装机编号",
    "产品",
    "型号",
    "数量",
    "配置",
    "销售",
    "销售渠道",
    "最终客户",
    "省份",
    "订单时间",
    "装机时间",
    "保修过期时间",
]

wb = Workbook()
wb.remove(wb.active)

header_fill = PatternFill("solid", fgColor="18274A")
body_fill = PatternFill("solid", fgColor="FAF8F7")
note_fill = PatternFill("solid", fgColor="E9E4E3")
header_font = Font(color="FFFFFF", bold=True)
body_font = Font(color="10203D")
thin = Side(style="thin", color="D4CAC8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_range(ws, min_row, max_row, max_col):
    for cell in ws[min_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.fill = body_fill
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border


def add_table(ws, table_name, max_row, max_col):
    table = Table(displayName=table_name, ref=f"A1:{ws.cell(row=max_row, column=max_col).coordinate}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


instructions = wb.create_sheet("说明")
instructions.merge_cells("A1:L1")
instructions["A1"] = "Baserow 装机数据导入模板"
instructions["A1"].fill = header_fill
instructions["A1"].font = Font(color="FFFFFF", bold=True, size=18)
instructions["A1"].alignment = Alignment(horizontal="center", vertical="center")
instructions.row_dimensions[1].height = 30

instruction_rows = [
    ["使用步骤", "只维护“装机数据”这一个工作表，每行是一条装机/销售明细。"],
    ["", "产品请选择 1180 Magnus OR Table、Tegris Dashboard、IC MIC Dashboard。"],
    ["", "装机编号用于增量更新：同一编号会更新旧记录；空白时系统会按产品、型号、客户、日期和渠道自动生成。"],
    ["", "型号、销售、销售渠道、最终客户如 Baserow 主数据中不存在，导入时会自动补充主数据并建立链接。"],
    ["", "导入后会增量写入 Baserow 的 Install_Base，并自动更新前台地图、排名、动态和趋势。"],
    ["", "1180 与 IC MIC 导入记录会在 Baserow Remarks 中自动标记为测试数据；Tegris 不会被标记。"],
    ["", "数量可以大于 1；日期建议使用 yyyy-mm-dd，例如 2026-04-20。"],
    ["", "页面右上角“导入装机数据”按钮在每个 dashboard 页面都可使用。"],
]
for row_index, row in enumerate(instruction_rows, start=3):
    instructions.cell(row=row_index, column=1, value=row[0])
    instructions.cell(row=row_index, column=2, value=row[1])
    instructions.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=12)
    instructions.row_dimensions[row_index].height = 28

for row in instructions.iter_rows(min_row=3, max_row=10, max_col=12):
    for cell in row:
        cell.fill = note_fill
        cell.font = body_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border

instructions.column_dimensions["A"].width = 14
for col in "BCDEFGHIJKL":
    instructions.column_dimensions[col].width = 16

config = wb.create_sheet("看板配置")
config_headers = ["产品", "productKey", "型号选项", "说明"]
config.append(config_headers)
for product_key, dashboard in dashboards.items():
    config.append([
        product_labels.get(product_key, product_key),
        product_key,
        "、".join(dashboard.get("productModels", [])),
        "这些型号会出现在模板下拉框中；也可以填 Baserow Product_Master 已有的其他型号。",
    ])
style_range(config, 1, config.max_row, len(config_headers))
for index, width in enumerate([28, 16, 42, 62], start=1):
    config.column_dimensions[config.cell(row=1, column=index).column_letter].width = width
add_table(config, "DashboardConfig", config.max_row, len(config_headers))
config.freeze_panes = "A2"

data = wb.create_sheet("装机数据")
data.append(headers)

while data.max_row < 101:
    data.append([""] * len(headers))

style_range(data, 1, data.max_row, len(headers))
for index, width in enumerate([18, 24, 22, 10, 30, 18, 30, 34, 16, 16, 16, 18], start=1):
    data.column_dimensions[data.cell(row=1, column=index).column_letter].width = width

product_validation = DataValidation(
    type="list",
    formula1=f'"{",".join(product_labels.values())}"',
    allow_blank=True,
)
data.add_data_validation(product_validation)
product_validation.add(f"B2:B{data.max_row}")

models = []
for dashboard in dashboards.values():
    for model in dashboard.get("productModels", []):
        if model not in models:
            models.append(model)
model_validation = DataValidation(type="list", formula1=f'"{",".join(models)}"', allow_blank=True)
data.add_data_validation(model_validation)
model_validation.add(f"C2:C{data.max_row}")

for col in ["J", "K", "L"]:
    for cell in data[col][1:]:
        cell.number_format = "yyyy-mm-dd"

add_table(data, "InstallBaseImport", data.max_row, len(headers))
data.freeze_panes = "A2"

wb.save(output_dir / "dashboard-data-import-template.xlsx")
