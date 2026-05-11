from __future__ import annotations

import json
import re
from copy import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path("/Users/wutiansun/Library/Mobile Documents/com~apple~CloudDocs/工作文件/副本1180用户名单（带调查需求）更新版_0401.xlsx")
TEMPLATE = ROOT / "dashboard-data-import-template.xlsx"
OUTPUT_JSON = ROOT / "outputs" / "1180-import" / "prepared-1180-records.json"

PROVINCE_SUFFIX = {
    "北京": "北京市",
    "天津": "天津市",
    "上海": "上海市",
    "重庆": "重庆市",
    "内蒙古": "内蒙古自治区",
    "广西": "广西壮族自治区",
    "宁夏": "宁夏回族自治区",
    "新疆": "新疆维吾尔自治区",
    "西藏": "西藏自治区",
}


def normalized_text(value) -> str:
    return str(value or "").strip()


def canonical_province(value) -> str:
    text = normalized_text(value)
    if not text:
        return ""
    if text in PROVINCE_SUFFIX:
        return PROVINCE_SUFFIX[text]
    if text.endswith(("省", "市", "自治区")):
        return text
    return f"{text}省"


def add_months(value: date, months: int) -> date:
    month_index = value.month - 1 + months
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    month_lengths = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    return date(year, month, min(value.day, month_lengths[month - 1]))


def install_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and int(value) == value and 1900 <= int(value) <= 2100:
        return date(int(value), 12, 31)

    text = normalized_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d{4}", text):
        return date(int(text), 12, 31)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m", "%Y/%m", "%Y.%m"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt in ("%Y-%m", "%Y/%m", "%Y.%m"):
                return date(parsed.year, parsed.month, 1)
            return parsed.date()
        except ValueError:
            pass
    return None


def stable_serial(row_number: int) -> str:
    return f"1180-0401-ROW{row_number:03d}"


def model_from_column_o(value) -> str:
    text = normalized_text(value).upper()
    suffix = text[-2:] if len(text) >= 2 else ""
    return suffix if suffix in {"B0", "B1", "B2", "B3", "B4", "B5"} else "缺省"


def load_records() -> list[dict]:
    wb = load_workbook(SOURCE, data_only=True)
    ws = wb["Hybrid OR Map"]
    records = []
    for row_number in range(5, ws.max_row + 1):
        sales = normalized_text(ws.cell(row_number, 2).value)
        province = canonical_province(ws.cell(row_number, 3).value)
        city = normalized_text(ws.cell(row_number, 4).value)
        raw_date = ws.cell(row_number, 5).value
        terminal_user = normalized_text(ws.cell(row_number, 6).value)
        product_model = model_from_column_o(ws.cell(row_number, 15).value)
        quantity = ws.cell(row_number, 16).value
        if not (sales and province and terminal_user):
            continue

        installed_on = install_date(raw_date)
        if not installed_on:
            continue

        try:
            quantity = int(quantity or 1)
        except (TypeError, ValueError):
            quantity = 1

        warranty_expires_on = add_months(installed_on, 24)
        records.append(
            {
                "sourceRow": row_number,
                "serialNo": stable_serial(row_number),
                "productKey": "magnus1180",
                "productModel": product_model,
                "quantity": quantity,
                "configDescription": "",
                "installProvince": province,
                "installCity": city,
                "terminalUser": terminal_user,
                "channelName": sales,
                "salesName": sales,
                "salesDate": installed_on.isoformat(),
                "installDate": installed_on.isoformat(),
                "warrantyExpireDate": warranty_expires_on.isoformat(),
            }
        )
    return records


def write_template(records: list[dict]) -> None:
    wb = load_workbook(TEMPLATE)
    ws = wb["1180用户列表"]
    headers = [
        "装机编号",
        "型号选择",
        "数量",
        "配置描述",
        "装机省份",
        "装机城市",
        "终端用户",
        "销售渠道",
        "销售",
        "销售时间",
        "装机时间",
        "保修过期时间",
    ]

    style_source = [ws.cell(2, min(col, ws.max_column)) for col in range(1, len(headers) + 1)]
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    for col, source_cell in enumerate(style_source, start=1):
        target = ws.cell(1, col)
        target._style = copy(source_cell._style)
        target.font = copy(source_cell.font)
        target.fill = copy(source_cell.fill)
        target.border = copy(source_cell.border)
        target.alignment = copy(source_cell.alignment)
        target.number_format = source_cell.number_format

    for record in records:
        ws.append(
            [
                record["serialNo"],
                record["productModel"],
                record["quantity"],
                record["configDescription"],
                record["installProvince"],
                record["installCity"],
                record["terminalUser"],
                record["channelName"],
                record["salesName"],
                record["salesDate"],
                record["installDate"],
                record["warrantyExpireDate"],
            ]
        )

    widths = [18, 12, 8, 16, 14, 14, 34, 18, 18, 14, 14, 16]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=10, max_col=12):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"

    wb.save(TEMPLATE)


def main() -> None:
    records = load_records()
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_template(records)
    total = sum(record["quantity"] for record in records)
    years = sorted({record["installDate"][:4] for record in records})
    print(json.dumps({"records": len(records), "totalQuantity": total, "firstYear": years[0], "lastYear": years[-1]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
