from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "副本SWP 1180 Funnel-2026-05-20-08-24-39.xlsx"
OUTPUT_DIR = ROOT / "outputs" / "magnus-2026-funnel-sync"
SOURCE_MARKER = "SOURCE_MAGNUS_2026_FUNNEL_20260520"
VALID_FUNNEL_MODELS = {"20", "40", "60", "80", "100"}

PROVINCE_SUFFIX = {
    "北京": "北京市",
    "天津": "天津市",
    "上海": "上海市",
    "重庆": "重庆市",
    "内蒙古": "内蒙古自治区",
    "广西": "广西壮族自治区",
    "宁夏": "宁夏回族自治区",
    "新疆": "新疆维吾尔自治区",
    "西藏": "西藏自治区",
}

PROVINCE_ALIASES = {
    "北京市": ["北京", "解放军总医院"],
    "天津市": ["天津"],
    "上海市": ["上海", "Shanghai", "复旦"],
    "重庆市": ["重庆"],
    "河北省": ["河北", "石家庄", "唐山", "秦皇岛", "邯郸", "邢台", "保定", "张家口", "承德", "沧州", "廊坊", "衡水"],
    "山西省": ["山西", "太原", "大同", "阳泉", "长治", "晋城", "朔州", "晋中", "运城", "忻州", "临汾", "吕梁"],
    "内蒙古自治区": ["内蒙古", "呼和浩特", "包头", "乌海", "赤峰", "通辽", "鄂尔多斯", "呼伦贝尔", "巴彦淖尔", "乌兰察布"],
    "辽宁省": ["辽宁", "沈阳", "大连", "鞍山", "抚顺", "本溪", "丹东", "锦州", "营口", "阜新", "辽阳", "盘锦", "铁岭", "朝阳", "葫芦岛", "中国医科大学", "盛京"],
    "吉林省": ["吉林", "长春", "四平", "辽源", "通化", "白山", "松原", "白城", "延边"],
    "黑龙江省": ["黑龙江", "哈尔滨", "齐齐哈尔", "鸡西", "鹤岗", "双鸭山", "大庆", "伊春", "佳木斯", "七台河", "牡丹江", "黑河", "绥化", "大兴安岭"],
    "江苏省": ["江苏", "南京", "无锡", "徐州", "常州", "苏州", "南通", "连云港", "淮安", "盐城", "扬州", "镇江", "泰州", "宿迁", "张家港", "滨海县", "靖江"],
    "浙江省": ["浙江", "杭州", "宁波", "温州", "嘉兴", "湖州", "绍兴", "金华", "衢州", "舟山", "台州", "丽水"],
    "安徽省": ["安徽", "合肥", "芜湖", "蚌埠", "淮南", "马鞍山", "淮北", "铜陵", "安庆", "黄山", "滁州", "阜阳", "宿州", "六安", "亳州", "池州", "宣城"],
    "福建省": ["福建", "福州", "厦门", "莆田", "三明", "泉州", "漳州", "南平", "龙岩", "宁德"],
    "江西省": ["江西", "南昌", "景德镇", "萍乡", "九江", "新余", "鹰潭", "赣州", "赣南", "吉安", "宜春", "抚州", "上饶"],
    "山东省": ["山东", "济南", "青岛", "淄博", "枣庄", "东营", "烟台", "潍坊", "济宁", "泰安", "威海", "日照", "临沂", "德州", "聊城", "滨州", "菏泽", "惠民县", "东阿", "即墨", "齐鲁"],
    "河南省": ["河南", "郑州", "开封", "洛阳", "平顶山", "安阳", "鹤壁", "新乡", "焦作", "濮阳", "许昌", "漯河", "三门峡", "南阳", "商丘", "信阳", "周口", "驻马店", "阜外华中心", "沁阳", "夏邑"],
    "湖北省": ["湖北", "武汉", "黄石", "十堰", "宜昌", "襄阳", "鄂州", "荆门", "孝感", "荆州", "黄冈", "咸宁", "随州", "恩施"],
    "湖南省": ["湖南", "长沙", "株洲", "湘潭", "衡阳", "邵阳", "岳阳", "常德", "张家界", "益阳", "郴州", "永州", "怀化", "娄底", "湘西", "湘乡"],
    "广东省": ["广东", "广州", "韶关", "深圳", "珠海", "汕头", "佛山", "江门", "湛江", "茂名", "肇庆", "惠州", "梅州", "汕尾", "河源", "阳江", "清远", "东莞", "中山", "潮州", "揭阳", "云浮"],
    "广西壮族自治区": ["广西", "南宁", "柳州", "桂林", "梧州", "北海", "防城港", "钦州", "贵港", "玉林", "百色", "贺州", "河池", "崇左"],
    "海南省": ["海南", "海口", "三亚", "三沙", "儋州"],
    "四川省": ["四川", "成都", "自贡", "攀枝花", "泸州", "德阳", "绵阳", "广元", "遂宁", "内江", "乐山", "南充", "眉山", "宜宾", "广安", "达州", "雅安", "巴中", "资阳", "阿坝", "甘孜", "凉山"],
    "贵州省": ["贵州", "贵阳", "六盘水", "遵义", "安顺", "毕节", "铜仁", "黔西南", "黔东南", "黔南"],
    "云南省": ["云南", "昆明", "曲靖", "玉溪", "保山", "昭通", "丽江", "普洱", "临沧", "楚雄", "红河", "文山", "西双版纳", "大理", "德宏", "怒江", "迪庆"],
    "西藏自治区": ["西藏", "拉萨", "日喀则", "昌都", "林芝", "山南", "那曲", "阿里"],
    "陕西省": ["陕西", "西安", "铜川", "宝鸡", "咸阳", "渭南", "延安", "汉中", "榆林", "安康", "商洛", "神木", "第九八六", "986医院"],
    "甘肃省": ["甘肃", "兰州", "嘉峪关", "金昌", "白银", "天水", "武威", "张掖", "平凉", "酒泉", "庆阳", "定西", "陇南", "临夏", "甘南"],
    "青海省": ["青海", "西宁", "海东", "海北", "黄南", "海南藏族自治州", "果洛", "玉树", "海西"],
    "宁夏回族自治区": ["宁夏", "银川", "石嘴山", "吴忠", "固原", "中卫"],
    "新疆维吾尔自治区": ["新疆", "乌鲁木齐", "克拉玛依", "吐鲁番", "哈密", "昌吉", "博尔塔拉", "巴音郭楞", "阿克苏", "克孜勒苏", "喀什", "和田", "伊犁", "塔城", "阿勒泰"],
}


def load_env() -> dict[str, str]:
    env = dict(os.environ)
    env_path = ROOT / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            env.setdefault(key.strip(), value.strip())
    return env


ENV = load_env()
API_URL = ENV.get("BASEROW_API_URL", "https://api.baserow.io").rstrip("/")
TOKEN = ENV.get("BASEROW_TOKEN", "")
TABLES = {
    "install": ENV.get("BASEROW_INSTALL_BASE_TABLE_ID", ""),
    "product": ENV.get("BASEROW_PRODUCT_TABLE_ID", ""),
    "customer": ENV.get("BASEROW_CUSTOMER_TABLE_ID", ""),
    "partner": ENV.get("BASEROW_SALES_PARTNER_TABLE_ID", ""),
}


def normalized(value) -> str:
    return str(value or "").strip()


def compact(value) -> str:
    return re.sub(r"\s+", "", normalized(value)).lower()


def normalize_model(value) -> str:
    return re.sub(r"[\s-]+", "", normalized(value).lower())


def iso_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalized(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    match = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
    return ""


def canonical_province(value) -> str:
    text = normalized(value)
    if not text:
        return ""
    if text in PROVINCE_SUFFIX:
        return PROVINCE_SUFFIX[text]
    if text.endswith(("省", "市", "自治区")):
        return text
    for province, aliases in PROVINCE_ALIASES.items():
        if text == province or text in aliases:
            return province
    return f"{text}省"


def infer_province(*values: str) -> str:
    text = " ".join(normalized(value) for value in values if normalized(value))
    if not text:
        return ""
    for province, aliases in PROVINCE_ALIASES.items():
        if province in text:
            return province
        for alias in aliases:
            if alias and alias in text:
                return province
    return ""


def request(path: str, method="GET", body=None):
    if not TOKEN or any(not value for value in TABLES.values()):
        raise RuntimeError("Missing Baserow token or table IDs in .env")
    data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
    last_error = None
    for attempt in range(1, 6):
        req = urllib.request.Request(
            f"{API_URL}{path}",
            data=data,
            method=method,
            headers={
                "Authorization": f"Token {TOKEN}",
                **({"Content-Type": "application/json"} if body is not None else {}),
            },
        )
        try:
            with urllib.request.urlopen(req, timeout=60) as response:
                if response.status == 204:
                    return None
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as error:
            detail = error.read().decode("utf-8", "replace")
            last_error = RuntimeError(f"Baserow request failed {error.code}: {detail[:500]}")
            if error.code not in {429, 500, 502, 503, 504} or attempt == 5:
                raise last_error from error
        except TimeoutError as error:
            last_error = error
            if attempt == 5:
                raise RuntimeError(f"Baserow request timed out after retries: {path}") from error
        time.sleep(2 * attempt)
    raise RuntimeError(f"Baserow request failed after retries: {last_error}")


def rows(table_id: str) -> list[dict]:
    output = []
    path = f"/api/database/rows/table/{table_id}/?user_field_names=true&size=200"
    while path:
        page = request(path)
        output.extend(page.get("results", []))
        path = ""
        if page.get("next"):
            parsed = urllib.parse.urlparse(page["next"])
            path = f"{parsed.path}?{parsed.query}"
    return output


def field_names(table_id: str) -> set[str]:
    return {field["name"] for field in request(f"/api/database/fields/table/{table_id}/")}


def create_row(table_id: str, payload: dict):
    return request(f"/api/database/rows/table/{table_id}/?user_field_names=true", "POST", payload)


def update_row(table_id: str, row_id: int, payload: dict):
    return request(f"/api/database/rows/table/{table_id}/{row_id}/?user_field_names=true", "PATCH", payload)


def delete_row(table_id: str, row_id: int):
    return request(f"/api/database/rows/table/{table_id}/{row_id}/", "DELETE")


def first_linked_id(value):
    return value[0].get("id") if isinstance(value, list) and value else None


def first_linked_name(value) -> str:
    if isinstance(value, list) and value:
        return normalized(value[0].get("value") or value[0].get("name"))
    if isinstance(value, dict):
        return normalized(value.get("value") or value.get("name"))
    return normalized(value)


def selected_value(value) -> str:
    return normalized(value.get("value") if isinstance(value, dict) else value)


def load_maps():
    product_rows, customer_rows, partner_rows = rows(TABLES["product"]), rows(TABLES["customer"]), rows(TABLES["partner"])
    return {
        "products": product_rows,
        "customers": customer_rows,
        "partners": partner_rows,
        "productsById": {row["id"]: row for row in product_rows},
        "productsByModel": {normalize_model(row.get("Product_Model")): row for row in product_rows},
        "customersByName": {compact(row.get("End_Customer")): row for row in customer_rows if compact(row.get("End_Customer"))},
        "partnersByName": {compact(row.get("Name")): row for row in partner_rows if compact(row.get("Name"))},
    }


def load_source_records(maps: dict | None = None) -> list[dict]:
    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    worksheet = workbook["SWP 1180 Funnel"]
    records = []
    customers_by_name = (maps or {}).get("customersByName", {})
    for row_number in range(18, worksheet.max_row + 1):
        opportunity_number = normalized(worksheet.cell(row_number, 13).value)
        customer_name = normalized(worksheet.cell(row_number, 16).value)
        if not opportunity_number or not customer_name:
            continue

        probability = normalized(worksheet.cell(row_number, 11).value)
        probability_match = re.search(r"20|40|60|80|100", probability)
        if not probability_match or probability_match.group(0) not in VALID_FUNNEL_MODELS:
            raise RuntimeError(f"Row {row_number} has unsupported probability: {probability!r}")

        order_date = iso_date(worksheet.cell(row_number, 2).value)
        if not order_date:
            raise RuntimeError(f"Row {row_number} has invalid order date: {worksheet.cell(row_number, 2).value!r}")

        quantity_raw = worksheet.cell(row_number, 20).value
        price_raw = worksheet.cell(row_number, 21).value
        try:
            quantity = int(quantity_raw or 1)
        except (TypeError, ValueError):
            quantity = 1
        try:
            total_price = float(price_raw or 0)
        except (TypeError, ValueError):
            total_price = 0.0

        existing_customer = customers_by_name.get(compact(customer_name), {})
        province = canonical_province(existing_customer.get("Province")) if existing_customer.get("Province") else ""
        province_source = "existing_customer" if province else ""
        if not province:
            province = infer_province(
                customer_name,
                worksheet.cell(row_number, 14).value,
                worksheet.cell(row_number, 22).value,
                worksheet.cell(row_number, 7).value,
            )
            province_source = "source_text" if province else ""

        opportunity_name = normalized(worksheet.cell(row_number, 14).value)
        stage = normalized(worksheet.cell(row_number, 12).value)
        sales_name = normalized(worksheet.cell(row_number, 8).value) or "未填写销售"
        partner_name = normalized(worksheet.cell(row_number, 15).value) or sales_name
        product_name = normalized(worksheet.cell(row_number, 18).value)
        price_text = str(int(total_price)) if total_price.is_integer() else str(total_price)
        config_parts = [
            opportunity_name,
            product_name,
            f"Opportunity {opportunity_number}",
            f"Total Price {price_text}" if total_price else "",
            f"Stage {stage}" if stage else "",
        ]

        records.append(
            {
                "sourceRow": row_number,
                "serialNo": opportunity_number,
                "productKey": "magnus2026Funnel",
                "productModel": f"{probability_match.group(0)} Funnel",
                "winRate": int(probability_match.group(0)),
                "quantity": quantity,
                "totalPrice": total_price,
                "configDescription": " | ".join(part for part in config_parts if part),
                "installProvince": province,
                "provinceSource": province_source,
                "installCity": "",
                "terminalUser": customer_name,
                "channelName": partner_name,
                "salesName": sales_name,
                "salesDate": order_date,
                "installDate": "",
                "warrantyExpireDate": "",
                "opportunityName": opportunity_name,
                "stage": stage,
                "forecastCategory": normalized(worksheet.cell(row_number, 10).value),
                "description": normalized(worksheet.cell(row_number, 22).value),
                "accountId": normalized(worksheet.cell(row_number, 27).value),
                "sapAccountNumber": normalized(worksheet.cell(row_number, 9).value),
            }
        )
    return records


def is_funnel_install_row(row: dict, maps: dict) -> bool:
    product = maps["productsById"].get(first_linked_id(row.get("Product_Model")), {})
    product_family = normalized(product.get("Product_Family") or row.get("Product_Family"))
    product_model = normalized(product.get("Product_Model") or first_linked_name(row.get("Product_Model")))
    remarks = normalized(row.get("Remarks"))
    serial = normalized(row.get("Serial_No"))
    text = " ".join([product_family, product_model, selected_value(row.get("Product_Line")), row.get("Product_Family") or ""]).lower()
    return (
        SOURCE_MARKER in remarks
        or serial.startswith("O-")
        and ("magnus funnel" in text or "funnel" in text)
    )


def find_or_create_product(record: dict, maps: dict, apply: bool):
    key = normalize_model(record["productModel"])
    if key in maps["productsByModel"]:
        return maps["productsByModel"][key]
    payload = {
        "Product_ID": f"MAGNUS-FUNNEL-{record['winRate']}",
        "Product_Line": "SWP / OT",
        "Product_Family": "Magnus Funnel",
        "Product_Model": record["productModel"],
        "Standard_Config": "",
        "Remarks": "Magnus 2026 Funnel dashboard model",
    }
    if not apply:
        return {"id": f"DRY-PRODUCT-{key}", **payload}
    row = create_row(TABLES["product"], payload)
    maps["productsByModel"][key] = row
    maps["productsById"][row["id"]] = row
    return row


def find_or_create_customer(record: dict, maps: dict, apply: bool):
    key = compact(record["terminalUser"])
    existing = maps["customersByName"].get(key)
    payload = {
        "End_Customer": record["terminalUser"],
        "Province": record["installProvince"],
        "City": record["installCity"],
        "Remarks": SOURCE_MARKER,
    }
    if existing:
        update_payload = {}
        if record["installProvince"] and not normalized(existing.get("Province")):
            update_payload["Province"] = record["installProvince"]
        if record["installCity"] and not normalized(existing.get("City")):
            update_payload["City"] = record["installCity"]
        if update_payload and apply:
            existing = update_row(TABLES["customer"], existing["id"], update_payload)
            maps["customersByName"][key] = existing
        return existing
    payload["Customer_ID"] = f"AUTO-FUNNEL-CUST-{int(time.time() * 1000)}"
    if not apply:
        return {"id": f"DRY-CUSTOMER-{len(maps['customersByName'])}", **payload}
    row = create_row(TABLES["customer"], payload)
    maps["customersByName"][key] = row
    return row


def find_or_create_partner(name: str, partner_type: str, province: str, maps: dict, apply: bool):
    key = compact(name)
    existing = maps["partnersByName"].get(key)
    if existing:
        return existing
    payload = {
        "Party_ID": f"AUTO-FUNNEL-PARTNER-{int(time.time() * 1000)}",
        "Name": name,
        "Type": partner_type,
        "Province": province,
        "Remarks": SOURCE_MARKER,
    }
    if not apply:
        return {"id": f"DRY-PARTNER-{len(maps['partnersByName'])}", **payload}
    row = create_row(TABLES["partner"], payload)
    maps["partnersByName"][key] = row
    return row


def install_payload(record: dict, product: dict, customer: dict, sales: dict, channel_partner: dict, allowed_fields: set[str]) -> dict:
    remarks = [
        SOURCE_MARKER,
        f"source row {record['sourceRow']}",
        f"Opportunity Name: {record['opportunityName']}",
        f"Total Price: {record['totalPrice']:.2f}",
        f"Win Rate: {record['winRate']}%",
        f"Forecast: {record['forecastCategory']}" if record["forecastCategory"] else "",
        f"SAP Account: {record['sapAccountNumber']}" if record["sapAccountNumber"] else "",
        f"Account ID: {record['accountId']}" if record["accountId"] else "",
        f"Description: {record['description']}" if record["description"] else "",
    ]
    payload = {
        "Serial_No": record["serialNo"],
        "Product_Family": "Magnus Funnel",
        "Product_Model": [product["id"]],
        "Product_Config": record["configDescription"],
        "Quantity": record["quantity"],
        "Sales": [sales["id"]],
        "End_Customer": [customer["id"]],
        "Channel_Partner": [channel_partner["id"]],
        "Order_Date": record["salesDate"],
        "Installation_Date": None,
        "Acceptance_Date": None,
        "Warranty_Months": "24",
        "Warranty_Expiry_Date": None,
        "Lifecycle_Status": "Ordered",
        "Project_Type": "Funnel",
        "Project_Source": "Dashboard Excel Import",
        "Remarks": " | ".join(part for part in remarks if part),
    }
    if "Funnel_Win_Rate" in allowed_fields:
        payload["Funnel_Win_Rate"] = record["winRate"]
    return {key: value for key, value in payload.items() if key in allowed_fields}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Write changes to Baserow. Omit for dry-run.")
    args = parser.parse_args()

    maps = load_maps()
    source_records = load_source_records(maps)
    unresolved = [record for record in source_records if not record["installProvince"]]
    if unresolved:
        sample = ", ".join(f"{record['sourceRow']}:{record['terminalUser']}" for record in unresolved[:12])
        raise RuntimeError(f"Could not infer province for {len(unresolved)} rows: {sample}")

    install_rows = rows(TABLES["install"])
    current_funnel_rows = [row for row in install_rows if is_funnel_install_row(row, maps)]
    current_serials = {normalized(row.get("Serial_No")) for row in current_funnel_rows}
    source_serials = {record["serialNo"] for record in source_records}
    products_to_create = sorted({
        record["productModel"]
        for record in source_records
        if normalize_model(record["productModel"]) not in maps["productsByModel"]
    })
    customers_to_create = sorted({
        record["terminalUser"]
        for record in source_records
        if compact(record["terminalUser"]) not in maps["customersByName"]
    })
    partners_to_create = sorted({
        record["salesName"]
        for record in source_records
        if compact(record["salesName"]) not in maps["partnersByName"]
    })

    summary = {
        "mode": "apply" if args.apply else "dry-run",
        "sourceRows": len(source_records),
        "sourceQuantity": sum(record["quantity"] for record in source_records),
        "sourceTotalPrice": sum(record["totalPrice"] for record in source_records),
        "sourceByWinRate": dict(Counter(str(record["winRate"]) for record in source_records)),
        "sourceByProvince": dict(Counter(record["installProvince"] for record in source_records)),
        "provinceSources": dict(Counter(record["provinceSource"] for record in source_records)),
        "currentBaserowFunnelRows": len(current_funnel_rows),
        "currentSerialsNotInSource": len(current_serials - source_serials),
        "sourceSerialsAlreadyInBaserow": len(source_serials & current_serials),
        "productsToCreate": products_to_create,
        "customersToCreate": len(customers_to_create),
        "partnersToCreate": len(partners_to_create),
        "deleteInstallRows": len(current_funnel_rows),
        "createInstallRows": len(source_records),
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "source-records-magnus-2026-funnel.json").write_text(
        json.dumps(source_records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (OUTPUT_DIR / "baserow-current-funnel-backup.json").write_text(
        json.dumps(current_funnel_rows, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    (OUTPUT_DIR / "sync-summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if not args.apply:
        return

    allowed_fields = field_names(TABLES["install"])
    for row in current_funnel_rows:
        delete_row(TABLES["install"], row["id"])

    created = 0
    for record in source_records:
        product = find_or_create_product(record, maps, True)
        customer = find_or_create_customer(record, maps, True)
        sales = find_or_create_partner(record["salesName"], "Sales", record["installProvince"], maps, True)
        channel_partner = find_or_create_partner(record["channelName"], "Channel Partner", record["installProvince"], maps, True)
        create_row(TABLES["install"], install_payload(record, product, customer, sales, channel_partner, allowed_fields))
        created += 1
        if created % 25 == 0:
            print(f"Created {created}/{len(source_records)} funnel rows", flush=True)

    final_summary = {
        **summary,
        "deletedInstallRowsApplied": len(current_funnel_rows),
        "createdInstallRowsApplied": created,
    }
    (OUTPUT_DIR / "sync-summary-applied.json").write_text(
        json.dumps(final_summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(final_summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
