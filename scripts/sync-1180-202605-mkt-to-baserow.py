from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "1180_202605IB list for MKT.xlsx"
OUTPUT_DIR = ROOT / "outputs" / "1180-202605-sync"
SOURCE_MARKER = "SOURCE_1180_202605_MKT"

PROVINCE_MAP = {
    "anhui": "安徽省",
    "beijing": "北京市",
    "北京直辖市": "北京市",
    "北京市": "北京市",
    "chong qing": "重庆市",
    "chongqing": "重庆市",
    "重庆直辖市": "重庆市",
    "重庆市": "重庆市",
    "fujian": "福建省",
    "gansu": "甘肃省",
    "guangdong": "广东省",
    "guangxi": "广西壮族自治区",
    "guizhou": "贵州省",
    "hainan": "海南省",
    "hebei": "河北省",
    "heilongjiang": "黑龙江省",
    "henan": "河南省",
    "hubei": "湖北省",
    "hunan": "湖南省",
    "jiangsu": "江苏省",
    "jiangxi": "江西省",
    "jilin": "吉林省",
    "liaoning": "辽宁省",
    "shaanxi": "陕西省",
    "shandong": "山东省",
    "shanghai": "上海市",
    "上海直辖市": "上海市",
    "上海市": "上海市",
    "shanxi": "山西省",
    "sichuan": "四川省",
    "tianjin": "天津市",
    "天津直辖市": "天津市",
    "yunnan": "云南省",
    "xinjiang": "新疆维吾尔自治区",
    "zhejiang": "浙江省",
}


def load_env() -> dict[str, str]:
    env = dict(os.environ)
    env_path = ROOT / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            env.setdefault(key.strip(), value.strip())
    return env


ENV = load_env()
API_URL = ENV.get("BASEROW_API_URL", "https://api.baserow.io").rstrip("/")
TOKEN = ENV.get("BASEROW_TOKEN", "")
TABLES = {
    "install": ENV.get("BASEROW_INSTALL_BASE_TABLE_ID", ""),
    "product": ENV.get("BASEROW_PRODUCT_TABLE_ID", ""),
    "customer": ENV.get("BASEROW_CUSTOMER_TABLE_ID", ""),
    "partner": ENV.get("BASEROW_SALES_PARTNER_TABLE_ID", ""),
}


def normalized(value) -> str:
    return str(value or "").strip()


def normalize_model(value) -> str:
    return re.sub(r"[\s-]+", "", normalized(value).lower())


def iso_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalized(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def canonical_province(value, fallback="") -> str:
    text = normalized(value)
    if not text:
        text = normalized(fallback)
    return PROVINCE_MAP.get(text.lower(), PROVINCE_MAP.get(text, text))


def model_from_material(value) -> str:
    text = normalized(value).upper()
    suffix = text[-2:] if len(text) >= 2 else ""
    return suffix if suffix in {"B0", "B1", "B2", "B3", "B4", "B5"} else "其他"


def load_source_records() -> list[dict]:
    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    worksheet = workbook["1180"]
    records = []
    for row_number in range(2, worksheet.max_row + 1):
        equipment_number = normalized(worksheet.cell(row_number, 1).value)
        if not equipment_number.isdigit():
            continue

        installed_on = iso_date(worksheet.cell(row_number, 11).value)
        material_number = normalized(worksheet.cell(row_number, 13).value)
        channel_name = normalized(worksheet.cell(row_number, 15).value)
        terminal_user = normalized(worksheet.cell(row_number, 18).value)
        province = canonical_province(worksheet.cell(row_number, 30).value, worksheet.cell(row_number, 17).value)
        if not (installed_on and material_number and channel_name and terminal_user and province):
            continue

        serial_number = normalized(worksheet.cell(row_number, 5).value)
        material_description = normalized(worksheet.cell(row_number, 3).value)
        warranty_expiry = iso_date(worksheet.cell(row_number, 7).value) or iso_date(worksheet.cell(row_number, 9).value)
        config_parts = [material_number, material_description, f"Column SN {serial_number}" if serial_number else ""]

        records.append({
            "sourceRow": row_number,
            "equipmentNumber": equipment_number,
            "serialNo": f"1180-202605-{equipment_number}",
            "productKey": "magnus1180",
            "productModel": model_from_material(material_number),
            "quantity": 1,
            "configDescription": " | ".join(part for part in config_parts if part),
            "installProvince": province,
            "installCity": "",
            "terminalUser": terminal_user,
            "channelName": channel_name,
            "salesName": channel_name,
            "salesDate": installed_on,
            "installDate": installed_on,
            "warrantyExpireDate": warranty_expiry,
            "gforceSystemId": equipment_number,
        })
    return records


def request(path: str, method="GET", body=None):
    if not TOKEN or any(not value for value in TABLES.values()):
        raise RuntimeError("Missing Baserow token or table IDs in .env")
    data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
    last_error = None
    for attempt in range(1, 6):
        req = urllib.request.Request(
            f"{API_URL}{path}",
            data=data,
            method=method,
            headers={
                "Authorization": f"Token {TOKEN}",
                **({"Content-Type": "application/json"} if body is not None else {}),
            },
        )
        try:
            with urllib.request.urlopen(req, timeout=60) as response:
                if response.status == 204:
                    return None
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as error:
            detail = error.read().decode("utf-8", "replace")
            last_error = RuntimeError(f"Baserow request failed {error.code}: {detail[:500]}")
            if error.code not in {429, 500, 502, 503, 504} or attempt == 5:
                raise last_error from error
        except TimeoutError as error:
            last_error = error
            if attempt == 5:
                raise RuntimeError(f"Baserow request timed out after retries: {path}") from error
        time.sleep(2 * attempt)
    raise RuntimeError(f"Baserow request failed after retries: {last_error}")


def rows(table_id: str) -> list[dict]:
    output = []
    path = f"/api/database/rows/table/{table_id}/?user_field_names=true&size=200"
    while path:
        page = request(path)
        output.extend(page.get("results", []))
        path = ""
        if page.get("next"):
            parsed = urllib.parse.urlparse(page["next"])
            path = f"{parsed.path}?{parsed.query}"
    return output


def field_names(table_id: str) -> set[str]:
    return {field["name"] for field in request(f"/api/database/fields/table/{table_id}/")}


def create_row(table_id: str, payload: dict):
    return request(f"/api/database/rows/table/{table_id}/?user_field_names=true", "POST", payload)


def update_row(table_id: str, row_id: int, payload: dict):
    return request(f"/api/database/rows/table/{table_id}/{row_id}/?user_field_names=true", "PATCH", payload)


def delete_row(table_id: str, row_id: int):
    return request(f"/api/database/rows/table/{table_id}/{row_id}/", "DELETE")


def first_linked_id(value):
    return value[0].get("id") if isinstance(value, list) and value else None


def first_linked_name(value) -> str:
    if isinstance(value, list) and value:
        return normalized(value[0].get("value") or value[0].get("name"))
    if isinstance(value, dict):
        return normalized(value.get("value") or value.get("name"))
    return normalized(value)


def selected_value(value) -> str:
    return normalized(value.get("value") if isinstance(value, dict) else value)


def load_maps():
    product_rows, customer_rows, partner_rows = rows(TABLES["product"]), rows(TABLES["customer"]), rows(TABLES["partner"])
    return {
        "products": product_rows,
        "customers": customer_rows,
        "partners": partner_rows,
        "productsById": {row["id"]: row for row in product_rows},
        "productsByModel": {normalize_model(row.get("Product_Model")): row for row in product_rows},
        "customersByName": {normalized(row.get("End_Customer")).lower(): row for row in customer_rows},
        "partnersByName": {normalized(row.get("Name")).lower(): row for row in partner_rows},
    }


def is_1180_install_row(row: dict, maps: dict) -> bool:
    product = maps["productsById"].get(first_linked_id(row.get("Product_Model")), {})
    product_family = normalized(product.get("Product_Family") or row.get("Product_Family"))
    product_model = normalized(product.get("Product_Model") or first_linked_name(row.get("Product_Model")))
    remarks = normalized(row.get("Remarks"))
    serial = normalized(row.get("Serial_No"))
    text = " ".join([product_family, product_model, selected_value(row.get("Product_Line"))]).lower()
    if "magnus funnel" in text or "funnel" in text:
        return False
    return (
        "TEST_DATA_1180" in remarks
        or serial.startswith("1180-202605-")
        or serial.startswith("1180-0401-")
        or (product_family.lower() == "magnus" and re.search(r"\bb[0-5]\b|缺省|其他", product_model.lower()))
    )


def find_or_create_product(record: dict, maps: dict, apply: bool):
    key = normalize_model(record["productModel"])
    if key in maps["productsByModel"]:
        return maps["productsByModel"][key]
    payload = {
        "Product_ID": f"AUTO-1180-{record['productModel']}-{int(time.time())}",
        "Product_Line": "SWP / OT",
        "Product_Family": "Magnus",
        "Product_Model": record["productModel"],
        "Standard_Config": "",
        "Remarks": SOURCE_MARKER,
    }
    if not apply:
        return {"id": f"DRY-PRODUCT-{key}", **payload}
    row = create_row(TABLES["product"], payload)
    maps["productsByModel"][key] = row
    maps["productsById"][row["id"]] = row
    return row


def find_or_create_customer(record: dict, maps: dict, apply: bool):
    key = normalized(record["terminalUser"]).lower()
    existing = maps["customersByName"].get(key)
    payload = {
        "End_Customer": record["terminalUser"],
        "Province": record["installProvince"],
        "City": record["installCity"],
        "Remarks": SOURCE_MARKER,
    }
    if existing:
        if apply and (
            normalized(existing.get("Province")) != record["installProvince"]
            or normalized(existing.get("City")) != record["installCity"]
        ):
            existing = update_row(TABLES["customer"], existing["id"], payload)
            maps["customersByName"][key] = existing
        return existing
    payload["Customer_ID"] = f"AUTO-CUST-{int(time.time() * 1000)}"
    if not apply:
        return {"id": f"DRY-CUSTOMER-{len(maps['customersByName'])}", **payload}
    row = create_row(TABLES["customer"], payload)
    maps["customersByName"][key] = row
    return row


def find_or_create_partner(name: str, province: str, maps: dict, apply: bool):
    key = normalized(name).lower()
    existing = maps["partnersByName"].get(key)
    if existing:
        return existing
    payload = {
        "Party_ID": f"AUTO-PARTNER-{int(time.time() * 1000)}",
        "Name": name,
        "Type": "Sales / Channel Partner",
        "Province": province,
        "Remarks": SOURCE_MARKER,
    }
    if not apply:
        return {"id": f"DRY-PARTNER-{len(maps['partnersByName'])}", **payload}
    row = create_row(TABLES["partner"], payload)
    maps["partnersByName"][key] = row
    return row


def install_payload(record: dict, product: dict, customer: dict, partner: dict, allowed_fields: set[str]) -> dict:
    payload = {
        "Serial_No": record["serialNo"],
        "Product_Family": "Magnus",
        "Product_Model": [product["id"]],
        "Product_Config": record["configDescription"],
        "Quantity": 1,
        "Sales": [partner["id"]],
        "End_Customer": [customer["id"]],
        "Channel_Partner": [partner["id"]],
        "City": record["installCity"],
        "Order_Date": record["salesDate"],
        "Installation_Date": record["installDate"],
        "Warranty_Expiry_Date": record["warrantyExpireDate"] or None,
        "Gforce_System_ID": record["gforceSystemId"],
        "Project_Source": "Dashboard Excel Import",
        "Remarks": f"TEST_DATA_1180 | {SOURCE_MARKER} | source row {record['sourceRow']}",
    }
    return {key: value for key, value in payload.items() if key in allowed_fields}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Write changes to Baserow. Omit for dry-run.")
    args = parser.parse_args()

    source_records = load_source_records()
    if len(source_records) != 415:
        raise RuntimeError(f"Expected 415 valid source rows, got {len(source_records)}")

    install_rows = rows(TABLES["install"])
    maps = load_maps()
    current_1180_rows = [row for row in install_rows if is_1180_install_row(row, maps)]
    current_serials = {normalized(row.get("Serial_No")) for row in current_1180_rows}
    source_serials = {record["serialNo"] for record in source_records}
    products_to_create = sorted({
        record["productModel"]
        for record in source_records
        if normalize_model(record["productModel"]) not in maps["productsByModel"]
    })
    customers_to_create = sorted({
        record["terminalUser"]
        for record in source_records
        if normalized(record["terminalUser"]).lower() not in maps["customersByName"]
    })
    partners_to_create = sorted({
        record["channelName"]
        for record in source_records
        if normalized(record["channelName"]).lower() not in maps["partnersByName"]
    })

    summary = {
        "mode": "apply" if args.apply else "dry-run",
        "sourceRows": len(source_records),
        "sourceQuantity": sum(record["quantity"] for record in source_records),
        "sourceByModel": dict(Counter(record["productModel"] for record in source_records)),
        "sourceByProvince": dict(Counter(record["installProvince"] for record in source_records)),
        "currentBaserow1180Rows": len(current_1180_rows),
        "currentBaserow1180Quantity": sum(int(row.get("Quantity") or 1) for row in current_1180_rows),
        "currentSerialsNotInSource": len(current_serials - source_serials),
        "sourceSerialsAlreadyInBaserow": len(source_serials & current_serials),
        "productsToCreate": products_to_create,
        "customersToCreate": len(customers_to_create),
        "partnersToCreate": len(partners_to_create),
        "deleteInstallRows": len(current_1180_rows),
        "createInstallRows": len(source_records),
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "source-records-1180-202605.json").write_text(
        json.dumps(source_records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (OUTPUT_DIR / "baserow-current-1180-backup.json").write_text(
        json.dumps(current_1180_rows, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    (OUTPUT_DIR / "sync-summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if not args.apply:
        return

    allowed_fields = field_names(TABLES["install"])
    for row in current_1180_rows:
        delete_row(TABLES["install"], row["id"])

    created = 0
    for record in source_records:
        product = find_or_create_product(record, maps, True)
        customer = find_or_create_customer(record, maps, True)
        partner = find_or_create_partner(record["channelName"], record["installProvince"], maps, True)
        create_row(TABLES["install"], install_payload(record, product, customer, partner, allowed_fields))
        created += 1
        if created % 50 == 0:
            print(f"Created {created}/{len(source_records)} install rows", flush=True)

    final_summary = {
        **summary,
        "deletedInstallRowsApplied": len(current_1180_rows),
        "createdInstallRowsApplied": created,
    }
    (OUTPUT_DIR / "sync-summary-applied.json").write_text(
        json.dumps(final_summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(final_summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
