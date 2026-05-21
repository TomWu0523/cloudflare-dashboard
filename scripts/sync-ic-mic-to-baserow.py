from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path("/Users/wutiansun/Documents/Cursor new project/IC IB 处理/dashboard-data-import-IC.xlsx")
OUTPUT_DIR = ROOT / "outputs" / "ic-mic-sync"
SOURCE_MARKER = "SOURCE_IC_MIC_20260521"
VALID_MODELS = {"HS66", "Novito", "S600"}

PROVINCE_MAP = {
    "北京": "北京市",
    "北京直辖市": "北京市",
    "北京市": "北京市",
    "上海": "上海市",
    "上海直辖市": "上海市",
    "上海市": "上海市",
    "天津": "天津市",
    "天津直辖市": "天津市",
    "天津市": "天津市",
    "重庆": "重庆市",
    "重庆直辖市": "重庆市",
    "重庆市": "重庆市",
    "内蒙古": "内蒙古自治区",
    "内蒙古自治区": "内蒙古自治区",
    "广西": "广西壮族自治区",
    "广西壮族自治区": "广西壮族自治区",
    "宁夏": "宁夏回族自治区",
    "宁夏回族自治区": "宁夏回族自治区",
    "新疆": "新疆维吾尔自治区",
    "新疆维吾尔自治区": "新疆维吾尔自治区",
    "西藏": "西藏自治区",
    "西藏自治区": "西藏自治区",
}

PROVINCE_ALIASES = {
    "北京市": ["北京", "解放军总医院"],
    "天津市": ["天津"],
    "上海市": ["上海"],
    "重庆市": ["重庆"],
    "河北省": ["河北", "石家庄", "唐山", "秦皇岛", "邯郸", "保定", "廊坊"],
    "山西省": ["山西", "太原", "大同", "长治", "晋城", "运城"],
    "内蒙古自治区": ["内蒙古", "呼和浩特", "包头", "鄂尔多斯"],
    "辽宁省": ["辽宁", "沈阳", "大连", "鞍山", "中国医科大学", "盛京"],
    "吉林省": ["吉林", "长春"],
    "黑龙江省": ["黑龙江", "哈尔滨", "齐齐哈尔", "大庆"],
    "江苏省": ["江苏", "南京", "无锡", "苏州", "南通", "徐州", "常州"],
    "浙江省": ["浙江", "杭州", "宁波", "温州", "嘉兴", "绍兴", "金华", "台州"],
    "安徽省": ["安徽", "合肥", "芜湖", "蚌埠", "安庆"],
    "福建省": ["福建", "福州", "厦门", "泉州", "漳州"],
    "江西省": ["江西", "南昌", "赣州"],
    "山东省": ["山东", "济南", "青岛", "烟台", "潍坊", "齐鲁"],
    "河南省": ["河南", "郑州", "洛阳", "南阳", "新乡"],
    "湖北省": ["湖北", "武汉", "宜昌", "襄阳"],
    "湖南省": ["湖南", "长沙", "湘雅", "衡阳"],
    "广东省": ["广东", "广州", "深圳", "东莞", "佛山", "珠海", "中山"],
    "广西壮族自治区": ["广西", "南宁", "柳州", "桂林"],
    "海南省": ["海南", "海口", "三亚"],
    "四川省": ["四川", "成都", "绵阳", "德阳"],
    "贵州省": ["贵州", "贵阳", "遵义"],
    "云南省": ["云南", "昆明"],
    "西藏自治区": ["西藏", "拉萨"],
    "陕西省": ["陕西", "西安", "西京", "空军军医"],
    "甘肃省": ["甘肃", "兰州"],
    "青海省": ["青海", "西宁"],
    "宁夏回族自治区": ["宁夏", "银川"],
    "新疆维吾尔自治区": ["新疆", "乌鲁木齐"],
}

PROVINCE_COORDINATES = {
    "上海市": [121.47, 31.23],
    "浙江省": [120.15, 30.28],
    "江苏省": [118.76, 32.06],
    "广东省": [113.27, 23.13],
    "四川省": [104.06, 30.67],
    "湖北省": [114.3, 30.59],
    "湖南省": [112.98, 28.19],
    "重庆市": [106.55, 29.56],
    "福建省": [119.3, 26.08],
    "安徽省": [117.28, 31.86],
    "广西壮族自治区": [108.32, 22.82],
    "云南省": [102.71, 25.04],
    "贵州省": [106.71, 26.58],
    "江西省": [115.86, 28.68],
    "海南省": [110.35, 20.02],
    "北京市": [116.41, 39.9],
    "天津市": [117.2, 39.12],
    "河北省": [114.48, 38.03],
    "山西省": [112.53, 37.87],
    "内蒙古自治区": [111.67, 40.82],
    "辽宁省": [123.43, 41.8],
    "吉林省": [125.32, 43.9],
    "黑龙江省": [126.63, 45.75],
    "山东省": [117, 36.65],
    "河南省": [113.62, 34.75],
    "西藏自治区": [91.13, 29.65],
    "陕西省": [108.94, 34.34],
    "甘肃省": [103.82, 36.06],
    "青海省": [101.78, 36.62],
    "宁夏回族自治区": [106.27, 38.47],
    "新疆维吾尔自治区": [87.62, 43.82],
}


def load_env() -> dict[str, str]:
    env = dict(os.environ)
    env_path = ROOT / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            env.setdefault(key.strip(), value.strip())
    return env


ENV = load_env()
API_URL = ENV.get("BASEROW_API_URL", "https://api.baserow.io").rstrip("/")
TOKEN = ENV.get("BASEROW_TOKEN", "")
TABLES = {
    "install": ENV.get("BASEROW_INSTALL_BASE_TABLE_ID", ""),
    "product": ENV.get("BASEROW_PRODUCT_TABLE_ID", ""),
    "customer": ENV.get("BASEROW_CUSTOMER_TABLE_ID", ""),
    "partner": ENV.get("BASEROW_SALES_PARTNER_TABLE_ID", ""),
}


def normalized(value) -> str:
    return str(value or "").strip()


def compact(value) -> str:
    return re.sub(r"\s+", "", normalized(value)).lower()


def normalize_model(value) -> str:
    return re.sub(r"[\s-]+", "", normalized(value).lower())


def iso_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalized(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def canonical_province(value) -> str:
    text = normalized(value)
    if not text:
        return ""
    if text in PROVINCE_MAP:
        return PROVINCE_MAP[text]
    if text.endswith(("省", "市", "自治区")):
        return text
    return f"{text}省"


def infer_province(*values) -> str:
    text = " ".join(normalized(value) for value in values if normalized(value))
    if not text:
        return ""
    for province, aliases in PROVINCE_ALIASES.items():
        if province in text:
            return province
        for alias in aliases:
            if alias and alias in text:
                return province
    return ""


def request(path: str, method="GET", body=None):
    if not TOKEN or any(not value for value in TABLES.values()):
        raise RuntimeError("Missing Baserow token or table IDs in .env")
    data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
    last_error = None
    for attempt in range(1, 6):
        req = urllib.request.Request(
            f"{API_URL}{path}",
            data=data,
            method=method,
            headers={
                "Authorization": f"Token {TOKEN}",
                **({"Content-Type": "application/json"} if body is not None else {}),
            },
        )
        try:
            with urllib.request.urlopen(req, timeout=60) as response:
                if response.status == 204:
                    return None
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as error:
            detail = error.read().decode("utf-8", "replace")
            last_error = RuntimeError(f"Baserow request failed {error.code}: {detail[:500]}")
            if error.code not in {429, 500, 502, 503, 504} or attempt == 5:
                raise last_error from error
        except TimeoutError as error:
            last_error = error
            if attempt == 5:
                raise RuntimeError(f"Baserow request timed out after retries: {path}") from error
        time.sleep(2 * attempt)
    raise RuntimeError(f"Baserow request failed after retries: {last_error}")


def rows(table_id: str) -> list[dict]:
    output = []
    path = f"/api/database/rows/table/{table_id}/?user_field_names=true&size=200"
    while path:
        page = request(path)
        output.extend(page.get("results", []))
        path = ""
        if page.get("next"):
            parsed = urllib.parse.urlparse(page["next"])
            path = f"{parsed.path}?{parsed.query}"
    return output


def field_names(table_id: str) -> set[str]:
    return {field["name"] for field in request(f"/api/database/fields/table/{table_id}/")}


def create_row(table_id: str, payload: dict):
    return request(f"/api/database/rows/table/{table_id}/?user_field_names=true", "POST", payload)


def update_row(table_id: str, row_id: int, payload: dict):
    return request(f"/api/database/rows/table/{table_id}/{row_id}/?user_field_names=true", "PATCH", payload)


def delete_row(table_id: str, row_id: int):
    return request(f"/api/database/rows/table/{table_id}/{row_id}/", "DELETE")


def chunks(items: list, size: int = 200):
    for index in range(0, len(items), size):
        yield items[index:index + size]


def create_rows_batch(table_id: str, payloads: list[dict]) -> list[dict]:
    created = []
    for payload_chunk in chunks(payloads):
        response = request(
            f"/api/database/rows/table/{table_id}/batch/?user_field_names=true",
            "POST",
            {"items": payload_chunk},
        )
        created.extend(response.get("items", []))
    return created


def delete_rows_batch(table_id: str, row_ids: list[int]):
    for id_chunk in chunks(row_ids):
        request(f"/api/database/rows/table/{table_id}/batch-delete/", "POST", {"items": id_chunk})


def first_linked_id(value):
    return value[0].get("id") if isinstance(value, list) and value else None


def first_linked_name(value) -> str:
    if isinstance(value, list) and value:
        return normalized(value[0].get("value") or value[0].get("name"))
    if isinstance(value, dict):
        return normalized(value.get("value") or value.get("name"))
    return normalized(value)


def selected_value(value) -> str:
    return normalized(value.get("value") if isinstance(value, dict) else value)


def load_maps():
    product_rows, customer_rows, partner_rows = rows(TABLES["product"]), rows(TABLES["customer"]), rows(TABLES["partner"])
    return {
        "products": product_rows,
        "customers": customer_rows,
        "partners": partner_rows,
        "productsById": {row["id"]: row for row in product_rows},
        "productsByModel": {normalize_model(row.get("Product_Model")): row for row in product_rows},
        "customersByName": {compact(row.get("End_Customer")): row for row in customer_rows if compact(row.get("End_Customer"))},
        "partnersByName": {compact(row.get("Name")): row for row in partner_rows if compact(row.get("Name"))},
    }


def load_source_records():
    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    worksheet = workbook["IC-MIC用户列表"]
    headers = [normalized(cell.value) for cell in worksheet[1]]
    skipped = []
    records = []
    for row_number in range(2, worksheet.max_row + 1):
        row = {header: worksheet.cell(row_number, index + 1).value for index, header in enumerate(headers) if header}
        if not any(normalized(value) for value in row.values()):
            continue
        model = normalized(row.get("型号选择"))
        quantity = int(row.get("数量") or 0)
        terminal_user = normalized(row.get("终端用户"))
        province = canonical_province(row.get("装机省份")) or infer_province(terminal_user)
        if not (model and quantity > 0 and terminal_user and province):
            skipped.append({"row": row_number, "model": model, "quantity": quantity, "terminalUser": terminal_user, "province": province})
            continue
        if model not in VALID_MODELS:
            raise RuntimeError(f"Row {row_number} has unsupported IC-MIC model: {model!r}")
        install_date = iso_date(row.get("装机时间"))
        sales_date = iso_date(row.get("销售时间")) or install_date
        records.append({
            "sourceRow": row_number,
            "serialNo": f"IC-MIC-20260521-ROW{row_number:04d}",
            "productKey": "icMic",
            "productModel": model,
            "quantity": quantity,
            "configDescription": normalized(row.get("配置描述")),
            "installProvince": province,
            "installCity": "",
            "terminalUser": terminal_user,
            "channelName": normalized(row.get("销售渠道")) or "未填写渠道",
            "salesName": normalized(row.get("销售")) or "未填写销售",
            "salesDate": sales_date,
            "installDate": install_date,
            "warrantyExpireDate": iso_date(row.get("保修过期时间")),
            "winRate": "",
            "gforceSystemId": "",
        })
    return records, skipped


def is_ic_install_row(row: dict, maps: dict) -> bool:
    product = maps["productsById"].get(first_linked_id(row.get("Product_Model")), {})
    product_line = selected_value(row.get("Product_Line"))
    product_family = normalized(product.get("Product_Family") or row.get("Product_Family"))
    product_model = normalized(product.get("Product_Model") or first_linked_name(row.get("Product_Model")))
    remarks = normalized(row.get("Remarks"))
    text = " ".join([product_line, product_family, product_model]).lower()
    return (
        SOURCE_MARKER in remarks
        or "TEST_DATA_IC_MIC" in remarks
        or "ic / mic" in text
        or "ic mic" in text
        or normalize_model(product_model) in {"hs66", "novito", "s600"}
    )


def find_or_create_product(record: dict, maps: dict, apply: bool):
    key = normalize_model(record["productModel"])
    if key in maps["productsByModel"]:
        return maps["productsByModel"][key]
    payload = {
        "Product_ID": f"IC-MIC-{record['productModel'].upper()}",
        "Product_Line": "IC / MIC",
        "Product_Family": "IC MIC",
        "Product_Model": record["productModel"],
        "Standard_Config": record["configDescription"],
        "Remarks": "IC MIC dashboard model",
    }
    if not apply:
        return {"id": f"DRY-PRODUCT-{key}", **payload}
    row = create_row(TABLES["product"], payload)
    maps["productsByModel"][key] = row
    maps["productsById"][row["id"]] = row
    return row


def find_or_create_customer(record: dict, maps: dict, apply: bool):
    key = compact(record["terminalUser"])
    existing = maps["customersByName"].get(key)
    payload = {
        "End_Customer": record["terminalUser"],
        "Province": record["installProvince"],
        "City": record["installCity"],
        "Remarks": SOURCE_MARKER,
    }
    if existing:
        update_payload = {}
        if record["installProvince"] and not normalized(existing.get("Province")):
            update_payload["Province"] = record["installProvince"]
        if record["installCity"] and not normalized(existing.get("City")):
            update_payload["City"] = record["installCity"]
        if update_payload and apply:
            existing = update_row(TABLES["customer"], existing["id"], update_payload)
            maps["customersByName"][key] = existing
        return existing
    payload["Customer_ID"] = f"AUTO-ICMIC-CUST-{int(time.time() * 1000)}"
    if not apply:
        return {"id": f"DRY-CUSTOMER-{len(maps['customersByName'])}", **payload}
    row = create_row(TABLES["customer"], payload)
    maps["customersByName"][key] = row
    return row


def find_or_create_partner(name: str, partner_type: str, province: str, maps: dict, apply: bool):
    key = compact(name)
    existing = maps["partnersByName"].get(key)
    if existing:
        return existing
    payload = {
        "Party_ID": f"AUTO-ICMIC-PARTNER-{int(time.time() * 1000)}",
        "Name": name,
        "Type": partner_type,
        "Province": province,
        "Remarks": SOURCE_MARKER,
    }
    if not apply:
        return {"id": f"DRY-PARTNER-{len(maps['partnersByName'])}", **payload}
    row = create_row(TABLES["partner"], payload)
    maps["partnersByName"][key] = row
    return row


def install_payload(record: dict, product: dict, customer: dict, sales: dict, channel_partner: dict, allowed_fields: set[str]) -> dict:
    payload = {
        "Serial_No": record["serialNo"],
        "Product_Family": "IC MIC",
        "Product_Model": [product["id"]],
        "Product_Config": record["configDescription"],
        "Quantity": record["quantity"],
        "Sales": [sales["id"]],
        "End_Customer": [customer["id"]],
        "Channel_Partner": [channel_partner["id"]],
        "City": record["installCity"] or None,
        "Order_Date": record["salesDate"] or None,
        "Installation_Date": record["installDate"] or None,
        "Warranty_Expiry_Date": record["warrantyExpireDate"] or None,
        "Gforce_System_ID": record["gforceSystemId"] or None,
        "Project_Source": "Dashboard Excel Import",
        "Remarks": f"TEST_DATA_IC_MIC | {SOURCE_MARKER} | source row {record['sourceRow']}",
    }
    return {key: value for key, value in payload.items() if key in allowed_fields}


def record_date(record: dict) -> str:
    return record.get("installDate") or record.get("salesDate") or ""


def dashboard_from_records(base: dict, records: list[dict]) -> dict:
    is_root_dashboard = bool(base.get("productLineOptions"))
    product_models = ["HS66", "S600", "Novito"]
    product_line_options = [
        {"value": "all", "label": "全部 IC MIC"},
        {"value": "hs66", "label": "HS66"},
        {"value": "s600", "label": "S600"},
        {"value": "novito", "label": "Novito"},
    ] if is_root_dashboard else []
    by_province = {}
    for record in records:
        province = record["installProvince"]
        current = by_province.setdefault(province, {
            "name": province,
            "value": 0,
            "latestSite": record["terminalUser"],
            "latestDate": record_date(record),
            "coord": PROVINCE_COORDINATES.get(province, [104.2, 35.8]),
            "_latest": "",
        })
        current["value"] += record["quantity"]
        current_date = record_date(record)
        if current_date >= current["_latest"]:
            current["latestSite"] = record["terminalUser"]
            current["latestDate"] = current_date
            current["_latest"] = current_date

    def top_group(key):
        grouped = {}
        for record in records:
            name = record[key]
            if not name:
                continue
            item = grouped.setdefault(name, {"name": name, "province": record["installProvince"], "value": 0})
            item["value"] += record["quantity"]
        return sorted(grouped.values(), key=lambda item: item["value"], reverse=True)[:8]

    dated = [record for record in records if record_date(record)]
    end_year = max((int(record_date(record)[:4]) for record in dated), default=datetime.now().year)
    years = [str(end_year - 4 + index) for index in range(5)]
    yearly = Counter(record_date(record)[:4] for record in dated)
    if dated:
        end = max(datetime.strptime(record_date(record), "%Y-%m-%d") for record in dated)
    else:
        end = datetime.now()
    months = []
    for offset in range(11, -1, -1):
        year = end.year + (end.month - offset - 1) // 12
        month = (end.month - offset - 1) % 12 + 1
        months.append(f"{year:04d}-{month:02d}")
    monthly = Counter(record_date(record)[:7] for record in dated)

    serializable_records = [{key: value for key, value in record.items() if key != "sourceRow"} for record in records]
    dashboard = {
        **base,
        "productModels": product_models,
        "productLineOptions": product_line_options,
        "provinceData": sorted(
            ({key: value for key, value in item.items() if key != "_latest"} for item in by_province.values()),
            key=lambda item: item["value"],
            reverse=True,
        ),
        "users": top_group("terminalUser"),
        "partners": top_group("channelName"),
        "updates": [
            {
                "date": (record_date(record) or "--")[5:] or "--",
                "status": "装机" if record.get("installDate") else "签约",
                "text": f"{record['terminalUser']} 完成 {record['productModel']} {record['quantity']} 台",
            }
            for record in sorted(records, key=record_date, reverse=True)[:6]
        ],
        "monthlyTrend": [{"month": month, "installed": monthly.get(month, 0)} for month in months],
        "yearlyTrend": [{"year": year, "installed": yearly.get(year, 0)} for year in years],
        "totalUnits": sum(record["quantity"] for record in records),
        "quarterUnits": 0,
        "sourceRecords": serializable_records,
    }
    if is_root_dashboard:
        dashboard["productLineData"] = {}
        for option in product_line_options:
            if option["value"] == "all":
                continue
            line_records = [record for record in records if normalize_model(record["productModel"]) == option["value"]]
            dashboard["productLineData"][option["value"]] = dashboard_from_records({**base, "productLineOptions": []}, line_records)
    return dashboard


def update_local_dashboard_data(records: list[dict]):
    data_file = ROOT / "data" / "dashboard-data.json"
    dashboards = json.loads(data_file.read_text(encoding="utf-8"))
    dashboards["icMic"] = dashboard_from_records(dashboards["icMic"], records)
    data_file.write_text(json.dumps(dashboards, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Write changes to Baserow. Omit for dry-run.")
    parser.add_argument("--update-local", action="store_true", help="Update local data/dashboard-data.json with IC-MIC records.")
    args = parser.parse_args()

    source_records, skipped = load_source_records()
    maps = load_maps()
    install_rows = rows(TABLES["install"])
    current_ic_rows = [row for row in install_rows if is_ic_install_row(row, maps)]
    current_serials = {normalized(row.get("Serial_No")) for row in current_ic_rows}
    source_serials = {record["serialNo"] for record in source_records}
    products_to_create = sorted({
        record["productModel"]
        for record in source_records
        if normalize_model(record["productModel"]) not in maps["productsByModel"]
    })
    customers_to_create = sorted({
        record["terminalUser"]
        for record in source_records
        if compact(record["terminalUser"]) not in maps["customersByName"]
    })
    partners_to_create = sorted({
        name
        for record in source_records
        for name in (record["salesName"], record["channelName"])
        if compact(name) not in maps["partnersByName"]
    })
    summary = {
        "mode": "apply" if args.apply else "dry-run",
        "sourceRows": len(source_records),
        "sourceQuantity": sum(record["quantity"] for record in source_records),
        "skippedRows": skipped,
        "sourceByModelRows": dict(Counter(record["productModel"] for record in source_records)),
        "sourceByModelQuantity": dict(Counter({model: sum(record["quantity"] for record in source_records if record["productModel"] == model) for model in VALID_MODELS})),
        "sourceByProvinceQuantity": dict(Counter({province: sum(record["quantity"] for record in source_records if record["installProvince"] == province) for province in sorted({record["installProvince"] for record in source_records})})),
        "currentBaserowIcRows": len(current_ic_rows),
        "currentBaserowIcQuantity": sum(int(row.get("Quantity") or 1) for row in current_ic_rows),
        "currentSerialsNotInSource": len(current_serials - source_serials),
        "sourceSerialsAlreadyInBaserow": len(source_serials & current_serials),
        "productsToCreate": products_to_create,
        "customersToCreate": len(customers_to_create),
        "partnersToCreate": len(partners_to_create),
        "deleteInstallRows": len(current_ic_rows),
        "createInstallRows": len(source_records),
    }
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "source-records-ic-mic.json").write_text(json.dumps(source_records, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "baserow-current-ic-mic-backup.json").write_text(json.dumps(current_ic_rows, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (OUTPUT_DIR / "sync-summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if args.update_local:
        update_local_dashboard_data(source_records)

    if not args.apply:
        return

    allowed_fields = field_names(TABLES["install"])
    if current_ic_rows:
        delete_rows_batch(TABLES["install"], [row["id"] for row in current_ic_rows])
        print(f"Deleted {len(current_ic_rows)} existing IC-MIC install rows", flush=True)

    product_payloads = []
    for model in sorted({record["productModel"] for record in source_records}):
        if normalize_model(model) not in maps["productsByModel"]:
            product_payloads.append({
                "Product_ID": f"IC-MIC-{model.upper()}",
                "Product_Line": "IC / MIC",
                "Product_Family": "IC MIC",
                "Product_Model": model,
                "Standard_Config": "",
                "Remarks": "IC MIC dashboard model",
            })
    for row in create_rows_batch(TABLES["product"], product_payloads):
        maps["productsByModel"][normalize_model(row.get("Product_Model"))] = row
        maps["productsById"][row["id"]] = row
    if product_payloads:
        print(f"Created {len(product_payloads)} IC-MIC product rows", flush=True)

    customer_payloads_by_key = {}
    for record in source_records:
        key = compact(record["terminalUser"])
        existing = maps["customersByName"].get(key)
        if existing:
            if record["installProvince"] and not normalized(existing.get("Province")):
                updated = update_row(TABLES["customer"], existing["id"], {
                    "Province": record["installProvince"],
                    "City": record["installCity"],
                })
                maps["customersByName"][key] = updated
            continue
        customer_payloads_by_key[key] = {
            "Customer_ID": f"AUTO-ICMIC-CUST-{int(time.time() * 1000)}-{len(customer_payloads_by_key) + 1}",
            "End_Customer": record["terminalUser"],
            "Province": record["installProvince"],
            "City": record["installCity"],
            "Remarks": SOURCE_MARKER,
        }
    for row in create_rows_batch(TABLES["customer"], list(customer_payloads_by_key.values())):
        maps["customersByName"][compact(row.get("End_Customer"))] = row
    if customer_payloads_by_key:
        print(f"Created {len(customer_payloads_by_key)} customer rows", flush=True)

    partner_payloads_by_key = {}
    for record in source_records:
        for name, partner_type in ((record["salesName"], "Sales"), (record["channelName"], "Channel Partner")):
            key = compact(name)
            if key in maps["partnersByName"] or key in partner_payloads_by_key:
                continue
            partner_payloads_by_key[key] = {
                "Party_ID": f"AUTO-ICMIC-PARTNER-{int(time.time() * 1000)}-{len(partner_payloads_by_key) + 1}",
                "Name": name,
                "Type": partner_type,
                "Province": record["installProvince"],
                "Remarks": SOURCE_MARKER,
            }
    for row in create_rows_batch(TABLES["partner"], list(partner_payloads_by_key.values())):
        maps["partnersByName"][compact(row.get("Name"))] = row
    if partner_payloads_by_key:
        print(f"Created {len(partner_payloads_by_key)} partner rows", flush=True)

    install_payloads = []
    for record in source_records:
        product = maps["productsByModel"][normalize_model(record["productModel"])]
        customer = maps["customersByName"][compact(record["terminalUser"])]
        sales = maps["partnersByName"][compact(record["salesName"])]
        channel = maps["partnersByName"][compact(record["channelName"])]
        install_payloads.append(install_payload(record, product, customer, sales, channel, allowed_fields))
    created = len(create_rows_batch(TABLES["install"], install_payloads))
    print(f"Created {created}/{len(source_records)} install rows", flush=True)

    final_summary = {
        **summary,
        "deletedInstallRowsApplied": len(current_ic_rows),
        "createdInstallRowsApplied": created,
    }
    (OUTPUT_DIR / "sync-summary-applied.json").write_text(json.dumps(final_summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(final_summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
